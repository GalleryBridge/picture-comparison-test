"""
报告生成器模块

基于比对结果生成详细的Excel和HTML格式报告。
提供完整的统计分析、差异详情和可视化图表。
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.drawing.image import Image
import jinja2
import os
import base64
import io
from typing import List, Dict, Tuple, Optional, Any
from dataclasses import dataclass, asdict
from enum import Enum
from pathlib import Path
import time

from geometry.elements import Element, Point, Line, Circle, Arc, Text
from matching.diff_detector import DifferenceDetail, DifferenceType, DifferenceStatistics
from core.comparison_engine import ComparisonResult
from visualization.diff_renderer import DiffRenderer, RenderConfig, RenderFormat


class ReportFormat(Enum):
    """报告格式"""
    EXCEL = "excel"
    HTML = "html"
    BOTH = "both"


class ReportLevel(Enum):
    """报告详细程度"""
    SUMMARY = "summary"       # 摘要报告
    DETAILED = "detailed"     # 详细报告
    COMPREHENSIVE = "comprehensive"  # 综合报告


@dataclass
class ReportConfig:
    """报告配置"""
    # 基本信息
    title: str = "PDF图纸比对报告"
    author: str = "PDF Comparison System"
    company: str = ""
    
    # 报告设置
    level: ReportLevel = ReportLevel.DETAILED
    include_charts: bool = True
    include_images: bool = True
    include_raw_data: bool = False
    
    # 样式配置
    theme_color: str = "FF2E8B57"  # 主题色 (aRGB格式)
    header_color: str = "FF4169E1"  # 标题色 (aRGB格式)
    highlight_color: str = "FFFFD700"  # 高亮色 (aRGB格式)
    
    # 文件配置
    image_dpi: int = 300
    image_format: RenderFormat = RenderFormat.PNG


class ReportGenerator:
    """报告生成器"""
    
    def __init__(self, config: Optional[ReportConfig] = None):
        self.config = config or ReportConfig()
        self.diff_renderer = DiffRenderer()
        
        # 设置Jinja2模板环境
        self.template_env = jinja2.Environment(
            loader=jinja2.FileSystemLoader('templates'),
            autoescape=jinja2.select_autoescape(['html', 'xml'])
        )
    
    def generate_report(self, comparison_result: ComparisonResult, 
                       output_path: str, format: ReportFormat = ReportFormat.EXCEL) -> bool:
        """生成报告"""
        
        try:
            if format == ReportFormat.EXCEL:
                return self._generate_excel_report(comparison_result, output_path)
            elif format == ReportFormat.HTML:
                return self._generate_html_report(comparison_result, output_path)
            elif format == ReportFormat.BOTH:
                # 生成两种格式
                excel_path = output_path.replace('.xlsx', '_excel.xlsx').replace('.html', '.xlsx')
                html_path = output_path.replace('.xlsx', '.html').replace('_excel.xlsx', '.html')
                
                excel_success = self._generate_excel_report(comparison_result, excel_path)
                html_success = self._generate_html_report(comparison_result, html_path)
                
                return excel_success and html_success
            
            return False
            
        except Exception as e:
            print(f"报告生成失败: {e}")
            return False
    
    def _generate_excel_report(self, comparison_result: ComparisonResult, output_path: str) -> bool:
        """生成Excel报告"""
        
        try:
            # 创建工作簿
            wb = openpyxl.Workbook()
            
            # 删除默认工作表
            wb.remove(wb.active)
            
            # 1. 生成摘要工作表
            self._create_summary_sheet(wb, comparison_result)
            
            # 2. 生成差异详情工作表
            self._create_differences_sheet(wb, comparison_result)
            
            # 3. 生成统计图表工作表
            if self.config.include_charts:
                self._create_charts_sheet(wb, comparison_result)
            
            # 4. 生成原始数据工作表
            if self.config.include_raw_data:
                self._create_raw_data_sheet(wb, comparison_result)
            
            # 保存工作簿
            wb.save(output_path)
            return True
            
        except Exception as e:
            print(f"Excel报告生成失败: {e}")
            return False
    
    def _create_summary_sheet(self, wb, comparison_result: ComparisonResult):
        """创建摘要工作表"""
        
        ws = wb.create_sheet("比对摘要")
        
        # 设置样式
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color=self.config.header_color, end_color=self.config.header_color, fill_type="solid")
        title_font = Font(bold=True, size=16, color=self.config.theme_color)
        
        # 标题
        ws['A1'] = self.config.title
        ws['A1'].font = title_font
        ws.merge_cells('A1:F1')
        
        # 基本信息
        row = 3
        basic_info = [
            ("比对ID", comparison_result.comparison_id),
            ("比对时间", comparison_result.timestamp),
            ("处理时间", f"{comparison_result.processing_time:.4f}秒"),
            ("文件A", comparison_result.file_a_path),
            ("文件B", comparison_result.file_b_path),
            ("", ""),
            ("图元数量", f"A: {comparison_result.elements_a_count}, B: {comparison_result.elements_b_count}"),
            ("匹配对数", comparison_result.matching_statistics.matched_pairs),
            ("平均相似度", f"{comparison_result.matching_statistics.average_similarity:.3f}"),
            ("", ""),
            ("差异统计", ""),
            ("总差异", comparison_result.difference_statistics.total_differences),
            ("新增", comparison_result.difference_statistics.added_count),
            ("删除", comparison_result.difference_statistics.deleted_count),
            ("修改", comparison_result.difference_statistics.modified_count),
            ("未变化", comparison_result.difference_statistics.unchanged_count),
            ("变化率", f"{comparison_result.difference_statistics.change_rate:.1%}")
        ]
        
        for label, value in basic_info:
            if label == "差异统计":
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = header_font
                ws[f'A{row}'].fill = header_fill
                row += 1
            elif label == "":
                row += 1
            else:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        
        # 添加边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(1, row):
            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
    
    def _create_differences_sheet(self, wb, comparison_result: ComparisonResult):
        """创建差异详情工作表"""
        
        ws = wb.create_sheet("差异详情")
        
        # 设置表头
        headers = ["序号", "差异类型", "描述", "相似度", "置信度", "图元A类型", "图元B类型", "几何变化", "属性变化"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.config.header_color, end_color=self.config.header_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # 填充数据
        for idx, diff in enumerate(comparison_result.differences, 2):
            ws.cell(row=idx, column=1, value=idx-1)  # 序号
            ws.cell(row=idx, column=2, value=diff.diff_type.value)  # 差异类型
            ws.cell(row=idx, column=3, value=diff.description)  # 描述
            ws.cell(row=idx, column=4, value=f"{diff.similarity:.3f}")  # 相似度
            ws.cell(row=idx, column=5, value=f"{diff.confidence:.3f}")  # 置信度
            
            # 图元类型
            elem_a_type = type(diff.element_a).__name__ if diff.element_a else "无"
            elem_b_type = type(diff.element_b).__name__ if diff.element_b else "无"
            ws.cell(row=idx, column=6, value=elem_a_type)
            ws.cell(row=idx, column=7, value=elem_b_type)
            
            # 变化详情
            geom_changes = str(diff.geometric_changes) if diff.geometric_changes else "无"
            attr_changes = str(diff.attribute_changes) if diff.attribute_changes else "无"
            ws.cell(row=idx, column=8, value=geom_changes)
            ws.cell(row=idx, column=9, value=attr_changes)
        
        # 调整列宽
        column_widths = [8, 12, 30, 10, 10, 12, 12, 20, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # 添加边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(1, len(comparison_result.differences) + 2):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _create_charts_sheet(self, wb, comparison_result: ComparisonResult):
        """创建图表工作表"""
        
        ws = wb.create_sheet("统计图表")
        
        # 1. 差异统计柱状图
        diff_data = [
            ["差异类型", "数量"],
            ["新增", comparison_result.difference_statistics.added_count],
            ["删除", comparison_result.difference_statistics.deleted_count],
            ["修改", comparison_result.difference_statistics.modified_count],
            ["未变化", comparison_result.difference_statistics.unchanged_count]
        ]
        
        for row, data in enumerate(diff_data, 1):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # 创建柱状图
        chart = BarChart()
        chart.title = "差异类型统计"
        chart.x_axis.title = "差异类型"
        chart.y_axis.title = "数量"
        
        data = Reference(ws, min_col=2, min_row=1, max_row=5, max_col=2)
        cats = Reference(ws, min_col=1, min_row=2, max_row=5)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "E2")
        
        # 2. 匹配统计饼图
        match_data = [
            ["匹配类型", "数量"],
            ["精确匹配", comparison_result.matching_statistics.exact_matches],
            ["相似匹配", comparison_result.matching_statistics.similar_matches],
            ["部分匹配", comparison_result.matching_statistics.partial_matches],
            ["未匹配", comparison_result.matching_statistics.unmatched_a + comparison_result.matching_statistics.unmatched_b]
        ]
        
        for row, data in enumerate(match_data, 8):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # 创建饼图
        pie_chart = PieChart()
        pie_chart.title = "匹配类型分布"
        
        data = Reference(ws, min_col=2, min_row=8, max_row=12, max_col=2)
        cats = Reference(ws, min_col=1, min_row=9, max_row=12)
        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(cats)
        
        ws.add_chart(pie_chart, "E15")
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
    
    def _create_raw_data_sheet(self, wb, comparison_result: ComparisonResult):
        """创建原始数据工作表"""
        
        ws = wb.create_sheet("原始数据")
        
        # 匹配统计原始数据
        ws['A1'] = "匹配统计原始数据"
        ws['A1'].font = Font(bold=True, size=14, color=self.config.theme_color)
        
        match_stats = asdict(comparison_result.matching_statistics)
        row = 3
        for key, value in match_stats.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            row += 1
        
        # 差异统计原始数据
        ws[f'A{row+1}'] = "差异统计原始数据"
        ws[f'A{row+1}'].font = Font(bold=True, size=14, color=self.config.theme_color)
        
        diff_stats = asdict(comparison_result.difference_statistics)
        row += 3
        for key, value in diff_stats.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def _generate_html_report(self, comparison_result: ComparisonResult, output_path: str) -> bool:
        """生成HTML报告"""
        
        try:
            # 生成图表图像
            chart_images = {}
            if self.config.include_images:
                chart_images = self._generate_chart_images(comparison_result)
            
            # 准备模板数据
            template_data = {
                'config': self.config,
                'result': comparison_result,
                'chart_images': chart_images,
                'generation_time': time.strftime('%Y-%m-%d %H:%M:%S')
            }
            
            # 渲染HTML模板
            html_content = self._render_html_template(template_data)
            
            # 保存HTML文件
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            return True
            
        except Exception as e:
            print(f"HTML报告生成失败: {e}")
            return False
    
    def _generate_chart_images(self, comparison_result: ComparisonResult) -> Dict[str, str]:
        """生成图表图像并转换为base64"""
        
        chart_images = {}
        
        try:
            with io.BytesIO() as buffer:
                # 生成比对摘要图表
                self.diff_renderer.render_comparison_summary(comparison_result, buffer, self.config.image_format)
                buffer.seek(0)
                chart_images['summary'] = base64.b64encode(buffer.getvalue()).decode()
            
            with io.BytesIO() as buffer:
                # 生成差异热力图
                self.diff_renderer.render_difference_heatmap(comparison_result, buffer, self.config.image_format)
                buffer.seek(0)
                chart_images['heatmap'] = base64.b64encode(buffer.getvalue()).decode()
            
            with io.BytesIO() as buffer:
                # 生成图元分布图
                self.diff_renderer.render_element_distribution(comparison_result, buffer, self.config.image_format)
                buffer.seek(0)
                chart_images['distribution'] = base64.b64encode(buffer.getvalue()).decode()
            
        except Exception as e:
            print(f"图表图像生成失败: {e}")
        
        return chart_images
    
    def _render_html_template(self, template_data: Dict[str, Any]) -> str:
        """渲染HTML模板"""
        
        # 内联HTML模板
        html_template = """
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ config.title }}</title>
    <style>
        body {
            font-family: 'Microsoft YaHei', Arial, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
            color: #333;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        .header {
            text-align: center;
            border-bottom: 3px solid {{ config.theme_color }};
            padding-bottom: 20px;
            margin-bottom: 30px;
        }
        .header h1 {
            color: {{ config.theme_color }};
            margin: 0;
            font-size: 28px;
        }
        .header p {
            color: #666;
            margin: 10px 0 0 0;
        }
        .section {
            margin-bottom: 30px;
        }
        .section h2 {
            color: {{ config.header_color }};
            border-left: 4px solid {{ config.header_color }};
            padding-left: 15px;
            margin-bottom: 20px;
        }
        .info-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 20px;
            margin-bottom: 20px;
        }
        .info-card {
            background-color: #f8f9fa;
            padding: 15px;
            border-radius: 8px;
            border-left: 4px solid {{ config.theme_color }};
        }
        .info-card h3 {
            margin: 0 0 10px 0;
            color: {{ config.theme_color }};
        }
        .info-card p {
            margin: 5px 0;
        }
        .stats-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin-bottom: 20px;
        }
        .stat-card {
            background: linear-gradient(135deg, {{ config.theme_color }}, {{ config.header_color }});
            color: white;
            padding: 20px;
            border-radius: 10px;
            text-align: center;
        }
        .stat-card h3 {
            margin: 0 0 10px 0;
            font-size: 24px;
        }
        .stat-card p {
            margin: 0;
            opacity: 0.9;
        }
        .chart-container {
            text-align: center;
            margin: 20px 0;
        }
        .chart-container img {
            max-width: 100%;
            height: auto;
            border: 1px solid #ddd;
            border-radius: 8px;
        }
        .differences-table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }
        .differences-table th,
        .differences-table td {
            border: 1px solid #ddd;
            padding: 12px;
            text-align: left;
        }
        .differences-table th {
            background-color: {{ config.header_color }};
            color: white;
            font-weight: bold;
        }
        .differences-table tr:nth-child(even) {
            background-color: #f8f9fa;
        }
        .diff-type-added { color: #28a745; font-weight: bold; }
        .diff-type-deleted { color: #dc3545; font-weight: bold; }
        .diff-type-modified { color: #007bff; font-weight: bold; }
        .diff-type-unchanged { color: #6c757d; font-weight: bold; }
        .footer {
            text-align: center;
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid #ddd;
            color: #666;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>{{ config.title }}</h1>
            <p>生成时间: {{ generation_time }}</p>
            <p>比对ID: {{ result.comparison_id }}</p>
        </div>

        <div class="section">
            <h2>基本信息</h2>
            <div class="info-grid">
                <div class="info-card">
                    <h3>文件信息</h3>
                    <p><strong>文件A:</strong> {{ result.file_a_path }}</p>
                    <p><strong>文件B:</strong> {{ result.file_b_path }}</p>
                    <p><strong>图元数量:</strong> A={{ result.elements_a_count }}, B={{ result.elements_b_count }}</p>
                </div>
                <div class="info-card">
                    <h3>比对信息</h3>
                    <p><strong>比对时间:</strong> {{ result.timestamp }}</p>
                    <p><strong>处理时间:</strong> {{ "%.4f"|format(result.processing_time) }}秒</p>
                    <p><strong>状态:</strong> {{ "成功" if result.success else "失败" }}</p>
                </div>
            </div>
        </div>

        <div class="section">
            <h2>统计概览</h2>
            <div class="stats-grid">
                <div class="stat-card">
                    <h3>{{ result.matching_statistics.matched_pairs }}</h3>
                    <p>匹配对数</p>
                </div>
                <div class="stat-card">
                    <h3>{{ "%.1f"|format(result.matching_statistics.average_similarity * 100) }}%</h3>
                    <p>平均相似度</p>
                </div>
                <div class="stat-card">
                    <h3>{{ result.difference_statistics.total_differences }}</h3>
                    <p>总差异数</p>
                </div>
                <div class="stat-card">
                    <h3>{{ "%.1f"|format(result.difference_statistics.change_rate * 100) }}%</h3>
                    <p>变化率</p>
                </div>
            </div>
        </div>

        {% if chart_images %}
        <div class="section">
            <h2>可视化图表</h2>
            {% if chart_images.summary %}
            <div class="chart-container">
                <h3>比对摘要</h3>
                <img src="data:image/png;base64,{{ chart_images.summary }}" alt="比对摘要图表">
            </div>
            {% endif %}
            {% if chart_images.heatmap %}
            <div class="chart-container">
                <h3>差异热力图</h3>
                <img src="data:image/png;base64,{{ chart_images.heatmap }}" alt="差异热力图">
            </div>
            {% endif %}
            {% if chart_images.distribution %}
            <div class="chart-container">
                <h3>图元分布图</h3>
                <img src="data:image/png;base64,{{ chart_images.distribution }}" alt="图元分布图">
            </div>
            {% endif %}
        </div>
        {% endif %}

        <div class="section">
            <h2>差异详情</h2>
            {% if result.differences %}
            <table class="differences-table">
                <thead>
                    <tr>
                        <th>序号</th>
                        <th>差异类型</th>
                        <th>描述</th>
                        <th>相似度</th>
                        <th>置信度</th>
                        <th>图元A类型</th>
                        <th>图元B类型</th>
                    </tr>
                </thead>
                <tbody>
                    {% for diff in result.differences %}
                    <tr>
                        <td>{{ loop.index }}</td>
                        <td><span class="diff-type-{{ diff.diff_type.value }}">{{ diff.diff_type.value }}</span></td>
                        <td>{{ diff.description }}</td>
                        <td>{{ "%.3f"|format(diff.similarity) }}</td>
                        <td>{{ "%.3f"|format(diff.confidence) }}</td>
                        <td>{{ type(diff.element_a).__name__ if diff.element_a else "无" }}</td>
                        <td>{{ type(diff.element_b).__name__ if diff.element_b else "无" }}</td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
            {% else %}
            <p>无差异检测到。</p>
            {% endif %}
        </div>

        <div class="footer">
            <p>报告由 {{ config.author }} 生成 | {{ generation_time }}</p>
        </div>
    </div>
</body>
</html>
        """
        
        template = jinja2.Template(html_template)
        return template.render(**template_data)
    
    def update_config(self, new_config: ReportConfig):
        """更新报告配置"""
        self.config = new_config
